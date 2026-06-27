#!/usr/bin/env python3
"""project-to-openproject.py — project the docs/backlog/ spine into an OpenProject-importable
work-package table for the OpenProject "Excel synchronization".

See references/integrations/openproject.md. Highlights:
  - Only docs/backlog/ matters (Epics → Features → User Stories; Tasks with --with-tasks).
    Source of truth stays in docs/backlog/; the OpenProject project is a *projection*.
  - Columns: Type · ID · Subject · Priority · Description · Parent · Relations.
  - **ID** is blank — OpenProject assigns its own numeric id on import (theirs != ours).
    Our stable id lives at the START of the Subject ("<EP/F/US-id> <business title>").
  - **Hierarchy is automatic**: the Subject is indented with 4 spaces per depth level
    (Epic 0 · Feature 1 · User story 2 · Task 3). On upload, the OpenProject Excel-sync
    macro reads the indentation, nests the work packages, and auto-fills the Parent column.
    So the user does NOT create the parent/child relations by hand.
  - **Description**: a User Story's description IS its BDD (the Gherkin scenarios); a Feature
    carries its business-language description; an Epic its product vision.
  - **Relations** (follows/blocks/…) are emitted as an (empty) column, because they require
    OpenProject's numeric ids — only available AFTER the first import. To use them: import →
    let OpenProject assign ids → EXPORT the work packages back from OpenProject → fill the
    Relations column with "<type> <openproject-id>" → re-import. (Two-pass by design.)
  - Priority maps the *"Interpop"* scale: 🔴 Immediate · 🟠 High · 🟡 Normal · 🟢 Low.

Usage:
  project-to-openproject.py [--root docs] [--with-tasks] [--out-dir openproject] [--apply]

Safety: DRY-RUN by default (prints the table, writes nothing). --apply writes
  <out-dir>/openproject-backlog.csv   (always — stdlib only)
  <out-dir>/openproject-backlog.xlsx  (only if `openpyxl` is installed)
Never overwrites an existing file. This is the skill default; the user can always adapt.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

COLUMNS = ["Type", "ID", "Subject", "Priority", "Description", "Parent", "Relations"]
INDENT = "    "  # 4 spaces = one hierarchy level in the OpenProject Excel-sync template
DEPTH = {"Epic": 0, "Feature": 1, "User story": 2, "Task": 3}

_PRIO_EMOJI = {"🔴": "Immediate", "🟠": "High", "🟡": "Normal", "🟢": "Low"}
_PRIO_WORD = {
    "imediata": "Immediate", "immediate": "Immediate",
    "alta": "High", "high": "High",
    "normal": "Normal",
    "baixa": "Low", "low": "Low",
}
_EP_RE = re.compile(r"EP-?\d+(?:\.\d+)*", re.I)
_F_RE = re.compile(r"\bF-?\d+\b", re.I)
_US_RE = re.compile(r"US-?\d+\.\d+", re.I)
_T_RE = re.compile(r"\b(?:T-?\d+\.\d+\.\d+[a-z]?|TX-?\d+)\b", re.I)
_BUG_RE = re.compile(r"BUG-?\d+", re.I)
_QA_RE = re.compile(r"QA-?\d+", re.I)
_ISS_RE = re.compile(r"ISS-?\d+", re.I)
_SPK_RE = re.compile(r"SPK-?\d+", re.I)

# Structural buckets: directories under backlog/ that COLLAPSE into a ROOT Epic (depth 0) on export
# — siblings of the feature-front Epics. Entry: (dirname, Epic subject [pt-BR], default child kind,
# [nested sub-buckets...]). en-CA folder names map to pt-BR Epic titles here (the round-trip contract).
# `bugs/` is handled SEPARATELY (it is a TYPE parented to the violated feature, not an Epic bucket).
_BUCKETS = [
    ("improvements", "Melhorias", "Feature", []),                       # product enhancements → Features/US
    ("support-quality-investigation", "Atividades de Apoio, Qualidade e Investigação", None, [
        ("support", "Apoio", "Task", []),                            # cross-cutting TX → Tasks
        ("qa", "Q&A", "Task", []),                                   # quality activities → Tasks
        ("issues", "Issues", "Task", [                              # triage inbox → Tasks
            ("spikes", "Spikes", "Task", []),                        # time-boxed investigations → Tasks
        ]),
    ]),
]


def _norm(ident: str) -> str:
    ident = ident.upper()
    if ident.startswith("US"):
        return ident  # USNN.M (no hyphen, per conventions)
    return re.sub(r"^(EP|TX|BUG|QA|ISS|SPK|F|T)-?(?=\d)", lambda m: m.group(1) + "-", ident)


def _detect_priority(block: str) -> str:
    for emoji, name in _PRIO_EMOJI.items():
        if emoji in block:
            return name
    m = re.search(r"(priorid\w*|priority)[^\n]*", block.lower())
    scope = m.group(0) if m else ""
    for word, name in _PRIO_WORD.items():
        if re.search(r"\b" + re.escape(word) + r"\b", scope):
            return name
    return "Normal"


def _title(heading: str, ident: str) -> str:
    t = re.sub(r"^#{1,6}\s*", "", heading).strip()
    t = re.sub(r"^" + re.escape(ident) + r"\s*[—:\-]\s*", "", t, flags=re.I).strip()
    t = re.sub(r"^(EP|F|US|T|TX)-?\d+(?:\.\d+)*\s*[—:\-]\s*", "", t, flags=re.I).strip()
    return t


def _bucket_child_kind(stem: str, heading: str, default_kind: str) -> tuple[str, str]:
    """Classify a bucket child by the id it carries (BUG → Bug; QA/ISS/SPK/TX/T → Task;
    US → User story; F → Feature); fall back to the bucket's default kind + the file slug."""
    blob = f"{stem} {heading}"
    if (m := _BUG_RE.search(blob)):
        return "Bug", _norm(m.group(0))
    if (m := _SPK_RE.search(blob)):
        return "Task", _norm(m.group(0))
    if (m := _ISS_RE.search(blob)):
        return "Task", _norm(m.group(0))
    if (m := _QA_RE.search(blob)):
        return "Task", _norm(m.group(0))
    if (m := _T_RE.search(blob)):
        return "Task", _norm(m.group(0))
    if (m := _US_RE.search(blob)):
        return "User story", _norm(m.group(0))
    if (m := _F_RE.search(blob)):
        return "Feature", _norm(m.group(0))
    return default_kind or "Task", stem


def _heading_level(line: str) -> int:
    m = re.match(r"(#{1,6})\s", line)
    return len(m.group(1)) if m else 0


def _gherkin_in(lines: list[str]) -> str:
    """Concatenate the ```gherkin fenced blocks found in `lines` (the BDD)."""
    out, infence = [], False
    for ln in lines:
        s = ln.strip()
        if s.startswith("```gherkin"):
            infence = True
            continue
        if s.startswith("```") and infence:
            infence = False
            out.append("")  # blank line between scenarios
            continue
        if infence:
            out.append(ln.rstrip())
    return "\n".join(out).strip()


def _section_after(lines: list[str], pred) -> list[str]:
    """Lines under the first heading matching `pred(line)`, until the next same/higher heading."""
    out, lvl, collecting = [], 0, False
    for ln in lines:
        hl = _heading_level(ln)
        if not collecting:
            if hl and pred(ln):
                lvl, collecting = hl, True
            continue
        if hl and hl <= lvl:
            break
        out.append(ln)
    return out


def _first_paragraph(lines: list[str]) -> str:
    para = []
    for ln in lines:
        s = ln.strip()
        if not s or s.startswith(("#", ">", "|", "-", "*", "<!--")):
            if para:
                break
            continue
        para.append(s)
    return " ".join(para).strip()


def _row(kind, ident, title, priority="Normal", description="", depth=None):
    d = DEPTH.get(kind, 0) if depth is None else depth
    subject = INDENT * d + f"{ident} {title}".strip()
    return {"Type": kind, "ID": "", "Subject": subject,
            "Priority": priority, "Description": description, "Parent": "", "Relations": ""}


def _emit_bucket(rows, bdir, epic_subject, default_kind, subbuckets, depth):
    """Emit a bucket directory as an Epic (at `depth`) + its file children (depth+1),
    then recurse into nested sub-buckets as child Epics. README → Epic description."""
    if not bdir.is_dir():
        return
    readme = bdir / "README.md"
    rtext = readme.read_text(encoding="utf-8") if readme.is_file() else ""
    rlines = rtext.splitlines()
    vision = (_first_paragraph(_section_after(rlines, lambda ln: re.search(r"vis[ãa]o|descri[çc]|vision", ln, re.I)))
              or _first_paragraph(rlines))
    rows.append(_row("Epic", epic_subject, "", _detect_priority(rtext[:800]), vision, depth=depth))
    for cf in sorted(bdir.glob("*.md")):
        if cf.stem in ("README", "_TEMPLATE"):
            continue
        ctext = cf.read_text(encoding="utf-8")
        clines = ctext.splitlines()
        ch1 = next((ln for ln in clines if ln.startswith("# ")), cf.stem)
        kind, ident = _bucket_child_kind(cf.stem, ch1, default_kind)
        cdesc = (_first_paragraph(_section_after(clines, lambda ln: re.search(r"descri[çc]|vis[ãa]o|vision", ln, re.I)))
                 or _first_paragraph(clines))
        rows.append(_row(kind, ident, _title(ch1, ident), _detect_priority(ctext[:800]), cdesc, depth=depth + 1))
    for sub_dir, sub_subject, sub_kind, sub_nested in subbuckets:
        _emit_bucket(rows, bdir / sub_dir, sub_subject, sub_kind, sub_nested, depth + 1)


def collect(root: Path, with_tasks: bool) -> list[dict]:
    backlog = root / "backlog"
    if not backlog.is_dir():
        sys.exit(f"error: no '{backlog}' — run from a project with the docs/ spine "
                 f"(assets/scaffold-structure.sh).")
    rows: list[dict] = []

    # --- Epics ---
    epdir = backlog / "epics"
    for f in sorted(epdir.glob("*.md")) if epdir.is_dir() else []:
        if f.stem == "_TEMPLATE":
            continue
        text = f.read_text(encoding="utf-8")
        lines = text.splitlines()
        h1 = next((ln for ln in lines if ln.startswith("# ")), f.stem)
        m = _EP_RE.search(f.stem) or _EP_RE.search(h1)
        if not m:
            continue
        ident = _norm(m.group(0))
        vision = _first_paragraph(_section_after(lines, lambda ln: re.search(r"vis[ãa]o", ln, re.I)))
        rows.append(_row("Epic", ident, _title(h1, ident), _detect_priority(text[:800]), vision))

    # --- Features (+ their User Stories with BDD; Tasks optional) ---
    fdir = backlog / "features"
    for f in sorted(fdir.glob("*.md")) if fdir.is_dir() else []:
        if f.stem == "_TEMPLATE":
            continue
        text = f.read_text(encoding="utf-8")
        lines = text.splitlines()
        h1 = next((ln for ln in lines if ln.startswith("# ")), f.stem)
        mf = _F_RE.search(f.stem) or _F_RE.search(h1)
        if not mf:
            continue
        fid = _norm(mf.group(0))
        fdesc = _first_paragraph(_section_after(lines, lambda ln: re.search(r"descri[çc]", ln, re.I)))
        rows.append(_row("Feature", fid, _title(h1, fid), _detect_priority(text[:800]), fdesc))

        # User Stories: heading carrying a US id; description = its BDD (Gherkin)
        us_starts = [i for i, ln in enumerate(lines) if _heading_level(ln) and _US_RE.search(ln)]
        for k, i in enumerate(us_starts):
            uid = _norm(_US_RE.search(lines[i]).group(0))
            lvl = _heading_level(lines[i])
            j = len(lines)
            for nxt in range(i + 1, len(lines)):
                if _heading_level(lines[nxt]) and _heading_level(lines[nxt]) <= lvl:
                    j = nxt
                    break
            bdd = _gherkin_in(lines[i + 1:j])
            rows.append(_row("User story", uid, _title(lines[i], uid), "Normal", bdd))

        if with_tasks:
            for tid in sorted({_norm(m.group(0)) for m in _T_RE.finditer(text)}):
                rows.append(_row("Task", tid, "", "Normal", ""))

    # --- Structural buckets → ROOT Epics (depth 0) + nested child Epics + their children ---
    # backlog/improvements/ and backlog/support-quality-investigation/ are DIRECTORIES, not EP-NN
    # files. The umbrella nests child Epics (support/qa/issues, and spikes under issues); the
    # recursive emitter walks them. en-CA folder → pt-BR Epic title (see _BUCKETS).
    for dirname, epic_subject, default_kind, subbuckets in _BUCKETS:
        _emit_bucket(rows, backlog / dirname, epic_subject, default_kind, subbuckets, depth=0)

    # --- bugs/ → type "Bug" parented to the violated US/Feature (NOT an Epic bucket) ---
    # A bug inherits the violated Feature's Epic. depth 0 keeps it off the bucket indentation;
    # the Parent column carries the real parent (the REST adapter wires it as a true link).
    bugsdir = backlog / "bugs"
    for cf in sorted(bugsdir.glob("*.md")) if bugsdir.is_dir() else []:
        if cf.stem in ("README", "_TEMPLATE"):
            continue
        ctext = cf.read_text(encoding="utf-8")
        clines = ctext.splitlines()
        ch1 = next((ln for ln in clines if ln.startswith("# ")), cf.stem)
        mb = _BUG_RE.search(cf.stem) or _BUG_RE.search(ch1)
        bid = _norm(mb.group(0)) if mb else cf.stem
        pv = _US_RE.search(ctext) or _F_RE.search(ctext)   # parent = the violated US/Feature
        bdesc = (_first_paragraph(_section_after(clines, lambda ln: re.search(r"defect|defeito|descri", ln, re.I)))
                 or _first_paragraph(clines))
        r = _row("Bug", bid, _title(ch1, bid), _detect_priority(ctext[:800]), bdesc, depth=0)
        r["Parent"] = _norm(pv.group(0)) if pv else ""
        rows.append(r)
    return rows


def write_outputs(rows: list[dict], out_dir: Path, apply: bool) -> None:
    csv_path = out_dir / "openproject-backlog.csv"
    xlsx_path = out_dir / "openproject-backlog.xlsx"

    if not apply:
        print(f"\n  PLAN  {csv_path}  ({len(rows)} work packages)")
        print(f"  PLAN  {xlsx_path}  (if openpyxl present)")
        print("\nDry-run — nothing written. Re-run with --apply.")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    if csv_path.exists():
        print(f"  skip  {csv_path} (exists — not overwritten)")
    else:
        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=COLUMNS)
            w.writeheader()
            w.writerows(rows)
        print(f"  write {csv_path}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ModuleNotFoundError:
        print("  note  openpyxl not installed — CSV only. `uv add openpyxl` (or pip) for .xlsx.")
        return
    if xlsx_path.exists():
        print(f"  skip  {xlsx_path} (exists — not overwritten)")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = wrap  # Description wraps
    for col, width in zip("ABCDEFG", (12, 6, 64, 11, 60, 8, 18)):
        ws.column_dimensions[col].width = width
    wb.save(xlsx_path)
    print(f"  write {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Project docs/backlog/ into an OpenProject work-package table.")
    ap.add_argument("--root", default="docs", help="docs root (default: docs)")
    ap.add_argument("--out-dir", default="openproject", help="output dir (default: openproject)")
    ap.add_argument("--with-tasks", action="store_true", help="also emit T/TX tasks as work packages")
    ap.add_argument("--apply", action="store_true", help="write the files (default: dry-run)")
    args = ap.parse_args()

    rows = collect(Path(args.root), args.with_tasks)
    counts: dict[str, int] = {}
    for r in rows:
        counts[r["Type"]] = counts.get(r["Type"], 0) + 1
    print("== project-to-openproject.py ==")
    print(f"root={args.root}/  out-dir={args.out_dir}/  with-tasks={args.with_tasks}  apply={args.apply}")
    print("work packages: " + (", ".join(f"{k}={v}" for k, v in counts.items()) or "(none found)"))
    print("columns: " + " | ".join(COLUMNS))
    print("hierarchy: Subject indented 4 spaces/level → OpenProject nests + fills Parent on upload.")
    print("relations: emitted empty — need OpenProject ids (2nd pass: import → export back → fill → re-import).")
    write_outputs(rows, Path(args.out_dir), args.apply)


if __name__ == "__main__":
    main()
